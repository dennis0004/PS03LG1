
import time
import datetime
import os
import pickle
import numpy as np
import openpyxl
from python_calamine import CalamineWorkbook
from collections import defaultdict

from Agent2611gb import Agent2611gb
from LangGame11 import LangGame11

def mainTrain(interactionName, interactionStr, gameName, Nconcept, algoName, lrLG, lrPS, tauPS, epsilonPS, Nagent, simStart, simEnd, tStart, tEnd, dirPrefix, isSaveAll, isSaveLexicon):
    gameStr = '%s-nC%d' % (gameName, Nconcept)

    if algoName == 'Qlg3-g_Qps-b':
        algoStr = '%s-lrLG%.2f-lrPS%.2f-tauPS%d' % (algoName, lrLG, lrPS, tauPS)

    dirName1 = '%s/data/%s_%s_%s_Nagent%d'%(dirPrefix, interactionStr, gameStr, algoStr, Nagent)
    if not os.path.exists('data'):
        os.makedirs('data')
    if not os.path.exists(dirName1):
        os.makedirs(dirName1)

    np.random.seed()

    opLL = []
    for i in range(Nagent):
        opL = list(range(Nagent))
        opL.remove(i)
        opLL.append(opL)

    for sim in range(simStart, simEnd + 1):
        dirName = '%s/sim%04d' % (dirName1, sim)
        if not os.path.exists(dirName):
            os.makedirs(dirName)
        print(dirName)

        saveEvery = 100

        if not isSaveAll and isSaveLexicon:
            LexiconC2WT = []
            LexiconW2CT = []
        if isSaveAll:
            Nelement1 = Nconcept * Nconcept * 2
            Nelement2 = Nagent
            StatQmeanT = np.zeros((tEnd//saveEvery, Nelement1))
            StatQvarT = np.zeros((tEnd//saveEvery, Nelement1))
            StatPmeanT = np.zeros((tEnd//saveEvery, Nelement1))
            StatPvarT = np.zeros((tEnd//saveEvery, Nelement1))

            LexiconC2WT = []
            LexiconW2CT = []
            AgentRewardT = np.zeros((tEnd, Nagent), dtype=int)
            CountSelectedT = np.zeros((tEnd, Nagent), dtype=int)

            StatQps1meanT = np.zeros((tEnd // saveEvery, Nelement2))
            StatQps1varT = np.zeros((tEnd // saveEvery, Nelement2))
            StatPps1meanT = np.zeros((tEnd // saveEvery, Nelement2))
            StatPps1varT = np.zeros((tEnd // saveEvery, Nelement2))

        StatRewardSumRMT = np.zeros((tEnd//saveEvery, 1))
        StatC2WpctAlignRMT = np.zeros((tEnd // saveEvery, 1))
        StatW2CpctAlignRMT = np.zeros((tEnd // saveEvery, 1))
        StatPctSuccessRMT = np.zeros((tEnd // saveEvery, 1))
        StatSpecificityAvgT = np.zeros((tEnd // saveEvery, 1)) # measured by w2cL
        StatDLST = np.zeros((tEnd // saveEvery, 1))

        LexiconC2WmembersT = []
        LexiconC2WmembersCT = []

        LexiconW2CmembersT = []
        LexiconW2CmembersCT = []

        # GraphPS10TreeWidthT = np.zeros((tEnd // saveEvery, 1), dtype=int)
        # GraphPS20TreeWidthT = np.zeros((tEnd // saveEvery, 1), dtype=int)
        GraphPS1DegVarT = np.zeros((tEnd // saveEvery, 1))

        StatRewardSumT = np.zeros((tEnd, 1), dtype=int)
        AgentPST = np.zeros((tEnd, Nagent), dtype=int)
        StatC2WpctAlignT = np.zeros((tEnd // saveEvery, 1))
        StatW2CpctAlignT = np.zeros((tEnd // saveEvery, 1))

        if tStart != 0:
            filePath1 = '%s/sim%04d.xlsx' % (dirName, sim)
            filePath2 = '%s/sim%04d_T%d.xlsx' % (dirName, sim, tStart)
            if os.path.exists(filePath1):
                if os.path.exists(filePath2):
                    print('_T%d exist' % (tStart))
                    filePath3 = '%s/sim%04d_Tx.xlsx' % (dirName, sim)
                    os.rename(filePath1, filePath3)
                else:
                    os.rename(filePath1, filePath2)

            wb = CalamineWorkbook.from_path(filePath2)

            if not isSaveAll and isSaveLexicon:
                LexiconC2WT = wb.get_sheet_by_name('LexiconC2WT_every%d' % (saveEvery)).to_python()
                LexiconW2CT = wb.get_sheet_by_name('LexiconW2CT_every%d' % (saveEvery)).to_python()
            if isSaveAll:
                StatQmeanTinit = wb.get_sheet_by_name('StatQmeanT_every%d' % (saveEvery)).to_python()
                StatQmeanTinit = np.array(StatQmeanTinit)
                StatQvarTinit = wb.get_sheet_by_name('StatQvarT_every%d' % (saveEvery)).to_python()
                StatQvarTinit = np.array(StatQvarTinit)
                StatPmeanTinit = wb.get_sheet_by_name('StatPmeanT_every%d' % (saveEvery)).to_python()
                StatPmeanTinit = np.array(StatPmeanTinit)
                StatPvarTinit = wb.get_sheet_by_name('StatPvarT_every%d' % (saveEvery)).to_python()
                StatPvarTinit = np.array(StatPvarTinit)
                StatQmeanT[:tStart // saveEvery] = StatQmeanTinit[:tStart // saveEvery]
                StatQvarT[:tStart // saveEvery] = StatQvarTinit[:tStart // saveEvery]
                StatPmeanT[:tStart // saveEvery] = StatPmeanTinit[:tStart // saveEvery]
                StatPvarT[:tStart // saveEvery] = StatPvarTinit[:tStart // saveEvery]

                LexiconC2WT = wb.get_sheet_by_name('LexiconC2WT_every%d' % (saveEvery)).to_python()
                LexiconW2CT = wb.get_sheet_by_name('LexiconW2CT_every%d' % (saveEvery)).to_python()
                AgentRewardTinit = wb.get_sheet_by_name('AgentRewardT').to_python()
                AgentRewardTinit = np.array(AgentRewardTinit)
                AgentRewardT[:tStart] = AgentRewardTinit[:tStart]
                CountSelectedTinit = wb.get_sheet_by_name('CountSelectedT').to_python()
                CountSelectedTinit = np.array(CountSelectedTinit)
                CountSelectedT[:tStart] = CountSelectedTinit[:tStart]

                StatQps1meanTinit = wb.get_sheet_by_name('StatQps1meanT_every%d' % (saveEvery)).to_python()
                StatQps1meanTinit = np.array(StatQps1meanTinit)
                StatQps1varTinit = wb.get_sheet_by_name('StatQps1varT_every%d' % (saveEvery)).to_python()
                StatQps1varTinit = np.array(StatQps1varTinit)
                StatPps1meanTinit = wb.get_sheet_by_name('StatPps1meanT_every%d' % (saveEvery)).to_python()
                StatPps1meanTinit = np.array(StatPps1meanTinit)
                StatPps1varTinit = wb.get_sheet_by_name('StatPps1varT_every%d' % (saveEvery)).to_python()
                StatPps1varTinit = np.array(StatPps1varTinit)
                StatQps1meanT[:tStart // saveEvery] = StatQps1meanTinit[:tStart // saveEvery]
                StatQps1varT[:tStart // saveEvery] = StatQps1varTinit[:tStart // saveEvery]
                StatPps1meanT[:tStart // saveEvery] = StatPps1meanTinit[:tStart // saveEvery]
                StatPps1varT[:tStart // saveEvery] = StatPps1varTinit[:tStart // saveEvery]

            StatRewardSumRMTinit = wb.get_sheet_by_name('StatRewardSumRMT_every%d' % (saveEvery)).to_python()
            StatRewardSumRMTinit = np.array(StatRewardSumRMTinit)
            StatC2WpctAlignRMTinit = wb.get_sheet_by_name('StatC2WpctAlignRMT_every%d' % (saveEvery)).to_python()
            StatC2WpctAlignRMTinit = np.array(StatC2WpctAlignRMTinit)
            StatW2CpctAlignRMTinit = wb.get_sheet_by_name('StatW2CpctAlignRMT_every%d' % (saveEvery)).to_python()
            StatW2CpctAlignRMTinit = np.array(StatW2CpctAlignRMTinit)
            StatPctSuccessRMTinit = wb.get_sheet_by_name('StatPctSuccessRMT_every%d' % (saveEvery)).to_python()
            StatPctSuccessRMTinit = np.array(StatPctSuccessRMTinit)
            StatSpecificityAvgTinit = wb.get_sheet_by_name('StatSpecAvgT_every%d' % (saveEvery)).to_python()
            StatSpecificityAvgTinit = np.array(StatSpecificityAvgTinit)
            StatDLSTinit = wb.get_sheet_by_name('StatDLST_every%d' % (saveEvery)).to_python()
            StatDLSTinit = np.array(StatDLSTinit)

            LexiconC2WmembersT = wb.get_sheet_by_name('LexiconC2WmembersT_every%d' % (saveEvery)).to_python()
            LexiconC2WmembersCT = wb.get_sheet_by_name('LexiconC2WmembersCT_every%d' % (saveEvery)).to_python()

            LexiconW2CmembersT = wb.get_sheet_by_name('LexiconW2CmembersT_every%d' % (saveEvery)).to_python()
            LexiconW2CmembersCT = wb.get_sheet_by_name('LexiconW2CmembersCT_every%d' % (saveEvery)).to_python()

            # GraphPS10TreeWidthTinit = wb.get_sheet_by_name('GraphPS10TreeWidthT_every%d' % (saveEvery)).to_python()
            # GraphPS10TreeWidthTinit = np.array(GraphPS10TreeWidthTinit)
            # GraphPS20TreeWidthTinit = wb.get_sheet_by_name('GraphPS20TreeWidthT_every%d' % (saveEvery)).to_python()
            # GraphPS20TreeWidthTinit = np.array(GraphPS20TreeWidthTinit)
            GraphPS1DegVarTinit = wb.get_sheet_by_name('GraphPS1DegVarT_every%d' % (saveEvery)).to_python()
            GraphPS1DegVarTinit = np.array(GraphPS1DegVarTinit)

            StatRewardSumTinit = wb.get_sheet_by_name('StatRewardSumT').to_python()
            StatRewardSumTinit = np.array(StatRewardSumTinit)
            AgentPSTinit = wb.get_sheet_by_name('AgentPST').to_python()
            AgentPSTinit = np.array(AgentPSTinit)
            StatC2WpctAlignTinit = wb.get_sheet_by_name('StatC2WpctAlignT_every%d' % (saveEvery)).to_python()
            StatC2WpctAlignTinit = np.array(StatC2WpctAlignTinit)
            StatW2CpctAlignTinit = wb.get_sheet_by_name('StatW2CpctAlignT_every%d' % (saveEvery)).to_python()
            StatW2CpctAlignTinit = np.array(StatW2CpctAlignTinit)

            StatRewardSumRMT[:tStart//saveEvery] = StatRewardSumRMTinit[:tStart//saveEvery]
            StatC2WpctAlignRMT[:tStart // saveEvery] = StatC2WpctAlignRMTinit[:tStart // saveEvery]
            StatW2CpctAlignRMT[:tStart // saveEvery] = StatW2CpctAlignRMTinit[:tStart // saveEvery]
            StatPctSuccessRMT[:tStart // saveEvery] = StatPctSuccessRMTinit[:tStart // saveEvery]
            StatSpecificityAvgT[:tStart // saveEvery] = StatSpecificityAvgTinit[:tStart // saveEvery]
            StatDLST[:tStart // saveEvery] = StatDLSTinit[:tStart // saveEvery]

            # GraphPS10TreeWidthT[:tStart // saveEvery] = GraphPS10TreeWidthTinit[:tStart // saveEvery]
            # GraphPS20TreeWidthT[:tStart // saveEvery] = GraphPS20TreeWidthTinit[:tStart // saveEvery]
            GraphPS1DegVarT[:tStart // saveEvery] = GraphPS1DegVarTinit[:tStart // saveEvery]

            StatRewardSumT[:tStart] = StatRewardSumTinit[:tStart]
            AgentPST[:tStart] = AgentPSTinit[:tStart]
            StatC2WpctAlignT[:tStart // saveEvery] = StatC2WpctAlignTinit[:tStart // saveEvery]
            StatW2CpctAlignT[:tStart // saveEvery] = StatW2CpctAlignTinit[:tStart // saveEvery]

            print('sim %d Xlsx loaded' % (sim))
        else:
            pass

        if tStart != 0:
            filePath1 = '%s/Agents-sim%04d_T%d.pickle' % (dirName, sim, tStart)
            f01 = open(filePath1, 'rb')
            agents = pickle.load(f01)
            f01.close()

            print('sim %d Agents loaded' % (sim))
        else:
            agents = []
            for i in range(Nagent):
                agent = None
                if algoName == 'Qlg3-g_Qps-b':
                    agent = Agent2611gb(Nconcept, Nagent, opLL[i], lrLG, lrPS, tauPS, i, name='%s_%d' % (algoName, i), isDebug=False)
                agents.append(agent)

        game1 = None
        if gameName == 'LangGame11':
            game1 = LangGame11(Nconcept)
        c1 = game1.reset()

        for t in range(tStart + 1, tEnd + 1):
            # language game
            AgentReward = np.zeros(Nagent, dtype=int)
            AgentPS = np.zeros(Nagent, dtype=int)
            CountSelected = np.zeros(Nagent, dtype=int)
            for i in range(Nagent):
                # PS
                j = agents[i].getActionPS(0)

                # LG
                p1 = np.random.choice([i, j])
                p2 = j if p1 == i else i

                w1 = agents[p1].getActionLG(c1)
                c2 = agents[p2].getActionLG(w1 + 100)

                c1_, reward = game1.step(c2)

                # train LG
                agents[p1].train1lg(c1, w1, reward)
                agents[p1].train1lg(w1 + 100, c1, reward)
                agents[p2].train1lg(c2, w1, reward)
                agents[p2].train1lg(w1 + 100, c2, reward)

                if reward > 0: # win
                    for w_ in range(Nconcept):
                        if w_ != w1:
                            agents[p1].train1lg(c1, w_, -reward)
                            agents[p1].train1lg(w_ + 100, c1, -reward)
                    for c_ in range(Nconcept):
                        if c_ != c2:
                            agents[p2].train1lg(c_, w1, -reward)
                            agents[p2].train1lg(w1 + 100, c_, -reward)

                # train PS
                agents[i].train1ps(0, j, reward)

                c1 = c1_

                AgentReward[i] += reward
                AgentReward[j] += reward
                AgentPS[i] = j
                CountSelected[j] += 1

            # record
            if isSaveAll:
                AgentRewardT[t - 1] = AgentReward
                CountSelectedT[t - 1] = CountSelected
            StatRewardSum = np.sum(AgentReward)
            StatRewardSumT[t - 1, 0] = StatRewardSum
            AgentPST[t - 1] = AgentPS

            if t % saveEvery == 0:
                if isSaveAll:
                    Qsum = np.zeros(Nelement1)
                    Qsqsum = np.zeros(Nelement1)
                    Psum = np.zeros(Nelement1)
                    Psqsum = np.zeros(Nelement1)
                    Qps1sum = np.zeros(Nelement2)
                    Qps1sqsum = np.zeros(Nelement2)
                    Pps1sum = np.zeros(Nelement2)
                    Pps1sqsum = np.zeros(Nelement2)
                    for i in range(Nagent):
                        QlgAll = agents[i].getQlgAll()
                        PlgAll = agents[i].getPlgAll()
                        Qsum += QlgAll
                        Qsqsum += QlgAll ** 2
                        Psum += PlgAll
                        Psqsum += PlgAll ** 2
                        Qps1 = agents[i].getQps1(0)
                        Qps1sq = agents[i].getQps1(0) ** 2
                        Pps1 = agents[i].getPolicyPS1(0)
                        Pps1sq = agents[i].getPolicyPS1(0) ** 2
                        Qps1sum += Qps1
                        Qps1sqsum += Qps1sq
                        Pps1sum += Pps1
                        Pps1sqsum += Pps1sq
                    StatQmeanT[t // saveEvery - 1] = Qsum / Nagent
                    StatQvarT[t // saveEvery - 1] = Qsqsum / Nagent - (Qsum / Nagent) ** 2
                    StatPmeanT[t // saveEvery - 1] = Psum / Nagent
                    StatPvarT[t // saveEvery - 1] = Psqsum / Nagent - (Psum / Nagent) ** 2
                    StatQps1meanT[t//saveEvery - 1] = Qps1sum / (Nagent)
                    StatQps1varT[t//saveEvery - 1] = Qps1sqsum / (Nagent) - (Qps1sum / (Nagent)) ** 2
                    StatPps1meanT[t//saveEvery - 1] = Pps1sum / (Nagent)
                    StatPps1varT[t//saveEvery - 1] = Pps1sqsum / (Nagent) - (Pps1sum / (Nagent)) ** 2

                Pps1sum = np.zeros(Nagent)
                for i in range(Nagent):
                    Pps1sum += agents[i].getPolicyPS1(0)
                GraphPS1DegVar = np.var(Pps1sum)
                GraphPS1DegVarT[t // saveEvery - 1, 0] = GraphPS1DegVar

                # GraphPS10TreeWidth = calGraphPSTreeWidth(AgentPST[t-10:, :], Nagent)
                # GraphPS10TreeWidthT[t // saveEvery - 1, 0] = GraphPS10TreeWidth
                # GraphPS20TreeWidth = calGraphPSTreeWidth(AgentPST[t-20:, :], Nagent)
                # GraphPS20TreeWidthT[t // saveEvery - 1, 0] = GraphPS20TreeWidth

            if t % saveEvery == 0:
                c2wLL = []
                w2cLL = []
                for i in range(Nagent):
                    c2wLL.append(agents[i].getC2WArray())
                    w2cLL.append(agents[i].getW2CArray())
                StatRewardSumRM = 0
                StatC2WpctAlignRM = 0
                StatW2CpctAlignRM = 0
                StatPctSuccessRM = 0
                StatC2WpctAlign = 0
                StatW2CpctAlign = 0
                c01 = np.arange(Nconcept, dtype=int)
                StatSpecificityAvg = 0 # from w2cL
                LexiconC2WDict = defaultdict(list)
                LexiconW2CDict = defaultdict(list)
                for i in range(Nagent):
                    # PS (randomly choose opponent)
                    j = np.random.choice(opLL[i])

                    # LG
                    p1 = np.random.choice([i, j])
                    p2 = j if p1 == i else i

                    w1 = agents[p1].getActionLG(c1)
                    c2 = agents[p2].getActionLG(w1 + 100)

                    c1_, reward = game1.step(c2)

                    c1 = c1_

                    StatRewardSumRM += reward
                    StatRewardSumRM += reward

                    # other measurements
                    c2wLi = c2wLL[i]
                    c2wLj = c2wLL[j]
                    StatC2WpctAlignRM += np.sum(c2wLi == c2wLj) / Nconcept

                    w2cLi = w2cLL[i]
                    w2cLj = w2cLL[j]
                    StatW2CpctAlignRM += np.sum(w2cLi == w2cLj) / Nconcept

                    c02 = w2cLj[c2wLi]
                    c03 = w2cLi[c2wLj]
                    StatPctSuccessRM += (np.sum(c01==c02) + np.sum(c01==c03)) / (2*Nconcept)

                    j2 = AgentPS[i]
                    c2wLj2 = c2wLL[j2]
                    w2cLj2 = w2cLL[j2]
                    StatC2WpctAlign += np.sum(c2wLi == c2wLj2) / Nconcept
                    StatW2CpctAlign += np.sum(w2cLi == w2cLj2) / Nconcept

                    cLi, countLi = np.unique(w2cLi, return_counts=True)
                    StatSpecificityAvg += np.sum(1/countLi)/Nconcept

                    lexiconC2Wstr = np.array2string(c2wLi, separator=',').replace(' ', '')
                    LexiconC2WDict[lexiconC2Wstr].append(i)

                    lexiconW2Cstr = np.array2string(w2cLi, separator=',').replace(' ', '')
                    LexiconW2CDict[lexiconW2Cstr].append(i)

                StatC2WpctAlignRM /= Nagent
                StatW2CpctAlignRM /= Nagent
                StatPctSuccessRM /= Nagent
                StatSpecificityAvg /= Nagent
                StatRewardSumRMT[t//saveEvery - 1, 0] = StatRewardSumRM
                StatC2WpctAlignRMT[t // saveEvery - 1, 0] = StatC2WpctAlignRM
                StatW2CpctAlignRMT[t // saveEvery - 1, 0] = StatW2CpctAlignRM
                StatPctSuccessRMT[t // saveEvery - 1, 0] = StatPctSuccessRM
                StatSpecificityAvgT[t//saveEvery - 1, 0] = StatSpecificityAvg
                StatC2WpctAlignT[t // saveEvery - 1, 0] = StatC2WpctAlign / Nagent
                StatW2CpctAlignT[t // saveEvery - 1, 0] = StatW2CpctAlign / Nagent

                LexiconC2WFList = sorted(LexiconC2WDict.items(), key=lambda x: len(x[1]), reverse=True)
                LexiconC2WList, LexiconC2WmembersList = list(zip(*LexiconC2WFList))
                LexiconC2WmembersCList = [len(e) for e in LexiconC2WmembersList]
                LexiconC2WmembersList = [str(e).replace(' ', '') for e in LexiconC2WmembersList]
                if isSaveAll or isSaveLexicon:
                    LexiconC2WT.append(LexiconC2WList)
                LexiconC2WmembersT.append(LexiconC2WmembersList)
                LexiconC2WmembersCT.append(LexiconC2WmembersCList)

                LexiconW2CFList = sorted(LexiconW2CDict.items(), key=lambda x: len(x[1]), reverse=True)
                LexiconW2CList, LexiconW2CmembersList = list(zip(*LexiconW2CFList))
                LexiconW2CmembersCList = [len(e) for e in LexiconW2CmembersList]
                LexiconW2CmembersList = [str(e).replace(' ', '') for e in LexiconW2CmembersList]
                if isSaveAll or isSaveLexicon:
                    LexiconW2CT.append(LexiconW2CList)
                LexiconW2CmembersT.append(LexiconW2CmembersList)
                LexiconW2CmembersCT.append(LexiconW2CmembersCList)
                DLstr = LexiconW2CList[0]
                DLL = DLstr[1:-1].replace('\n', '').split(',')
                DLL = list(map(int, DLL))
                cLi, countLi = np.unique(DLL, return_counts=True)
                DLS = np.sum(1/countLi) / Nconcept
                StatDLST[t//saveEvery - 1, 0] = DLS

            if t % 5000 == 0 or t == tEnd: # multiple of saveEvery
                print('sim %d time %d completed: StatRewardSumRM %.2f; StatC2WpctAlignRM %.2f; StatW2CpctAlignRM %.2f; StatPctSuccessRM %.2f; StatAvgSpecificity %.2f; ACCc2w %d; ACCw2c %d;' % (sim, t, StatRewardSumRM, StatC2WpctAlignRM, StatW2CpctAlignRM, StatPctSuccessRM, StatSpecificityAvg, LexiconC2WmembersCList[0], LexiconW2CmembersCList[0]))

            if t == tEnd:
                filePath1 = '%s/Agents-sim%04d_T%d.pickle' % (dirName, sim, t)
                f01 = open(filePath1, 'wb')
                pickle.dump(agents, f01)
                f01.close()

            if t == tEnd:
                wb = openpyxl.Workbook(write_only=True)

                if not isSaveAll and isSaveLexicon:
                    ws = wb.create_sheet('LexiconC2WT_every%d' % (saveEvery))
                    for LexiconC2W in LexiconC2WT:
                        ws.append(LexiconC2W)
                    ws = wb.create_sheet('LexiconW2CT_every%d' % (saveEvery))
                    for LexiconW2C in LexiconW2CT:
                        ws.append(LexiconW2C)
                if isSaveAll:
                    ws = wb.create_sheet('StatQmeanT_every%d' % (saveEvery))
                    for StatQmean in StatQmeanT:
                        ws.append(list(StatQmean))
                    ws = wb.create_sheet('StatQvarT_every%d' % (saveEvery))
                    for StatQvar in StatQvarT:
                        ws.append(list(StatQvar))
                    ws = wb.create_sheet('StatPmeanT_every%d' % (saveEvery))
                    for StatPmean in StatPmeanT:
                        ws.append(list(StatPmean))
                    ws = wb.create_sheet('StatPvarT_every%d' % (saveEvery))
                    for StatPvar in StatPvarT:
                        ws.append(list(StatPvar))

                    ws = wb.create_sheet('LexiconC2WT_every%d' % (saveEvery))
                    for LexiconC2W in LexiconC2WT:
                        ws.append(LexiconC2W)
                    ws = wb.create_sheet('LexiconW2CT_every%d' % (saveEvery))
                    for LexiconW2C in LexiconW2CT:
                        ws.append(LexiconW2C)
                    ws = wb.create_sheet('AgentRewardT')
                    for AgentReward in AgentRewardT:
                        ws.append(list(AgentReward))
                    ws = wb.create_sheet('CountSelectedT')
                    for CountSelected in CountSelectedT:
                        ws.append(list(CountSelected))

                    ws = wb.create_sheet('StatQps1meanT_every%d' % (saveEvery))
                    for StatQpsmean in StatQps1meanT:
                        ws.append(list(StatQpsmean))
                    ws = wb.create_sheet('StatQps1varT_every%d' % (saveEvery))
                    for StatQpsvar in StatQps1varT:
                        ws.append(list(StatQpsvar))
                    ws = wb.create_sheet('StatPps1meanT_every%d' % (saveEvery))
                    for StatPpsmean in StatPps1meanT:
                        ws.append(list(StatPpsmean))
                    ws = wb.create_sheet('StatPps1varT_every%d' % (saveEvery))
                    for StatPpsvar in StatPps1varT:
                        ws.append(list(StatPpsvar))

                ws = wb.create_sheet('StatRewardSumRMT_every%d' % (saveEvery))
                for StatRewardSumRM in StatRewardSumRMT:
                    ws.append(list(StatRewardSumRM))
                ws = wb.create_sheet('StatC2WpctAlignRMT_every%d' % (saveEvery))
                for StatC2WpctAlignRM in StatC2WpctAlignRMT:
                    ws.append(list(StatC2WpctAlignRM))
                ws = wb.create_sheet('StatW2CpctAlignRMT_every%d' % (saveEvery))
                for StatW2CpctAlignRM in StatW2CpctAlignRMT:
                    ws.append(list(StatW2CpctAlignRM))
                ws = wb.create_sheet('StatPctSuccessRMT_every%d' % (saveEvery))
                for StatPctSuccessRM in StatPctSuccessRMT:
                    ws.append(list(StatPctSuccessRM))
                ws = wb.create_sheet('StatSpecAvgT_every%d' % (saveEvery))
                for StatSpecificityAvg in StatSpecificityAvgT:
                    ws.append(list(StatSpecificityAvg))
                ws = wb.create_sheet('StatDLST_every%d' % (saveEvery))
                for StatDLS in StatDLST:
                    ws.append(list(StatDLS))

                ws = wb.create_sheet('LexiconC2WmembersT_every%d' % (saveEvery))
                for LexiconC2Wmembers in LexiconC2WmembersT:
                    ws.append(LexiconC2Wmembers)
                ws = wb.create_sheet('LexiconC2WmembersCT_every%d' % (saveEvery))
                for LexiconC2WmembersC in LexiconC2WmembersCT:
                    ws.append(LexiconC2WmembersC)

                ws = wb.create_sheet('LexiconW2CmembersT_every%d' % (saveEvery))
                for LexiconW2Cmembers in LexiconW2CmembersT:
                    ws.append(LexiconW2Cmembers)
                ws = wb.create_sheet('LexiconW2CmembersCT_every%d' % (saveEvery))
                for LexiconW2CmembersC in LexiconW2CmembersCT:
                    ws.append(LexiconW2CmembersC)

                # ws = wb.create_sheet('GraphPS10TreeWidthT_every%d' % (saveEvery))
                # for GraphPS10TreeWidth in GraphPS10TreeWidthT:
                #     ws.append(list(GraphPS10TreeWidth))
                # ws = wb.create_sheet('GraphPS20TreeWidthT_every%d' % (saveEvery))
                # for GraphPS20TreeWidth in GraphPS20TreeWidthT:
                #     ws.append(list(GraphPS20TreeWidth))
                ws = wb.create_sheet('GraphPS1DegVarT_every%d' % (saveEvery))
                for GraphPS1DegVar in GraphPS1DegVarT:
                    ws.append(list(GraphPS1DegVar))

                ws = wb.create_sheet('StatRewardSumT')
                for StatRewardSum in StatRewardSumT:
                    ws.append(list(StatRewardSum))
                ws = wb.create_sheet('AgentPST')
                for AgentPS in AgentPST:
                    ws.append(list(AgentPS))
                ws = wb.create_sheet('StatC2WpctAlignT_every%d' % (saveEvery))
                for StatC2WpctAlign in StatC2WpctAlignT:
                    ws.append(list(StatC2WpctAlign))
                ws = wb.create_sheet('StatW2CpctAlignT_every%d' % (saveEvery))
                for StatW2CpctAlign in StatW2CpctAlignT:
                    ws.append(list(StatW2CpctAlign))

                wb.save('%s/sim%04d.xlsx' % (dirName, sim))
                wb.close()
        if tStart != 0:
            filePath2 = '%s/sim%04d_T%d.xlsx' % (dirName, sim, tStart)
            if os.path.exists(filePath2):
                os.remove(filePath2)
            else:
                print('sim %d _T%d not exist!' % (sim, tStart))
    print('completed', dirName1)
    return 'sim %d-%d' % (simStart, simEnd)


if __name__ == '__main__':
    # language game (LG) (speaker, listener)
    # PS (choose to play with anyone actively); rPS = reward
    # (for each player: choose one partner)
    # Q: change Q-values in "both directions"; if success, lower Q-values for other Q(c,w)

    dirPrefix = '.'

    simStart = 1
    simEnd = 1

    tStart = 0
    tEnd = 5000

    interactionName = 'seq-PS03'
    interactionStr = interactionName

    gameName = 'LangGame11'
    Nconcept = 100

    lrLG = 0.7
    lrPS = 0.05

    algoName = 'Qlg3-g_Qps-b'
    tauPS = -15
    epsilonPS = 1.0

    Nagent = 100

    isSaveAll = False
    isSaveLexicon = True

    t1 = time.time()

    mainTrain(interactionName, interactionStr, gameName, Nconcept, algoName, lrLG, lrPS, tauPS, epsilonPS, Nagent, simStart, simEnd, tStart, tEnd, dirPrefix, isSaveAll, isSaveLexicon)

    t2 = time.time()
    print('time: %.4f hours'%((t2-t1)/3600))
    print('done', datetime.datetime.now())




